import os
from zipfile import ZipFile

from openpyxl.reader.excel import load_workbook
import xlrd

from path import resources_path, zip_path
from pypdf import PdfReader

def test_archive():
    resources_files = os.listdir(resources_path)
    with ZipFile(zip_path, 'r') as zf:
        assert resources_files == ZipFile.namelist(zf)

def test_txt_file():
    with ZipFile(zip_path, 'r') as zf:
        with open(os.path.join(resources_path,'Hello.txt')) as txt:
            original_file = txt.read()
            assert original_file == zf.read('Hello.txt').decode()

def test_pdf_file():
    with ZipFile(zip_path, 'r') as zf:
        original_file = PdfReader(os.path.join(resources_path,'Python Testing with Pytest (Brian Okken).pdf'))
        archived_file = PdfReader(zf.open('Python Testing with Pytest (Brian Okken).pdf'))
        assert len(original_file.pages) == len(archived_file.pages)
        assert original_file.pages[3].extract_text() == archived_file.pages[3].extract_text()


def test_xlsx_file():
    with (ZipFile(zip_path, 'r') as zf):
        original_file = load_workbook(os.path.join(resources_path,'file_example_XLSX_50.xlsx'))
        archived_file = load_workbook(zf.open('file_example_XLSX_50.xlsx'))
        for worksheet in original_file.sheetnames:
            original_sheet = original_file[worksheet]
            archived_sheet = archived_file[worksheet]
        for row in range(1, original_sheet.max_row + 1):
            for col in range(1, original_sheet.max_column + 1):
                original_cell = original_sheet.cell(row, col)
                archived_cell = archived_sheet.cell(row, col)
                assert original_cell.value == archived_cell.value

def test_xls_file():
    with (ZipFile(zip_path, 'r') as zf):
        with zf.open('file_example_XLS_10.xls') as xls:
            temp = xls.read()
            archived_file = xlrd.open_workbook(file_contents=temp)
        original_file = xlrd.open_workbook(os.path.join(resources_path, 'file_example_XLS_10.xls'))
        assert original_file.sheet_names() == archived_file.sheet_names()
        original_sheet = original_file.sheet_by_index(0)
        archived_sheet = archived_file.sheet_by_index(0)
        assert original_sheet.nrows == archived_sheet.nrows
        assert original_sheet.ncols == archived_sheet.ncols
        for rownum in range(original_sheet.nrows):
            assert original_sheet.row_values(rownum) == archived_sheet.row_values(rownum)

